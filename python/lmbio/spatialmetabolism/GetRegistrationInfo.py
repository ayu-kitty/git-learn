#!/opt/conda/bin/python
import argparse
import os
import re
import shutil
import rpy2.robjects as robjects
from rpy2.robjects.packages import importr
import openpyxl
import sys
import pandas as pd
import argparse

# 命令行传参
parser = argparse.ArgumentParser()
# 添加命令行参数
parser.add_argument('--code', '-c', help='执行编号', required=True)
parser.add_argument('--AnalysisId', '-id', default = "none",help='项目编号')


args = parser.parse_args()
execution_code = args.code
AnalysisId = args.AnalysisId


# 导入lmbio包
meta = importr('lmbio')
# 初始化R
r = robjects.r
# 获取当前路径的相对路径名称
current_path = os.getcwd()
relative_path = os.path.basename(current_path)
# match_result = re.search(r'(DZLM|DLM|DQD|LM|ZLM|DOE\d+)', relative_path)
# 使用正则表达式匹配路径中的analysis_id
if AnalysisId == "none":
    match_result = re.search(r'(DZLM\d+)', relative_path)
    if match_result:
        matched_analysis_id = match_result.group(1)
        analysis_id = matched_analysis_id + "-aa"  # 添加后缀
        original_stdout = sys.stdout
        null_file = open("/dev/null", "w")
        sys.stdout = null_file
        # 调用GetAnalystInfo函数并传递参数
        analyst_info = r.GetAnalystInfo(analysis_id)
        sys.stdout = original_stdout
        null_file.close()
    else:
        analysis_id = AnalysisId + "-aa"
        print("No matching analysis_id found in the current path.")
else:
    analysis_id = AnalysisId + "-aa"
    original_stdout = sys.stdout
    null_file = open("/dev/null", "w")
    sys.stdout = null_file
    # 调用GetAnalystInfo函数并传递参数
    analyst_info = r.GetAnalystInfo(analysis_id)
    sys.stdout = original_stdout
    null_file.close()

# 查找以analysis_id命名的文件夹
folder_to_copy_from = os.path.join(current_path, analysis_id)
if os.path.exists(folder_to_copy_from):
    excel_file_path = os.path.join(folder_to_copy_from, "分析确认单.xlsx")
    # 读取原始Excel文件中的所有工作表
    all_sheets = pd.read_excel(excel_file_path, sheet_name=None)
    output_writer = pd.ExcelWriter('样品登记单.xlsx', engine='xlsxwriter')
    # 遍历所有工作表，将其内容写入新的工作簿中
    for sheet_name, sheet_df in all_sheets.items():
        # 如果是"样本实验信息"工作表，跳过第一行
        if sheet_name in ['样本实验信息', '分析基本信息']:
            sheet_df.to_excel(output_writer, sheet_name=sheet_name, index=False, header=False)
        else:
            sheet_df.to_excel(output_writer, sheet_name=sheet_name, index=False)
    output_writer._save()
    shutil.rmtree(folder_to_copy_from)
else:
        print("Folder not found for copying.")

def get_value_from_sheet(sheet, key):
    for row in range(1, sheet.max_row + 1):
        cell_key = sheet.cell(row=row, column=1).value
        cell_value = sheet.cell(row=row, column=2).value
        if cell_key == key:
            return cell_value
    return ''

def remove_info(project_sheet):
        for row in range(2, project_sheet.max_row + 1):
            project_sheet.delete_rows(2)

def copy_info(online_group_sheet, project_group_sheet):
    for online_row in online_group_sheet.iter_rows(min_row=2, values_only=True):
        group = online_row[0]
        sample_names = online_row[1]
        if ',' in sample_names:
            samples = online_row[1].split(",")
            for row in range(1, len(samples)):
                for sample in samples:
                    project_group_sheet.append([group, sample])
        else:
            project_group_sheet.append([group, sample_names])

# 加载线上登记文件
if os.path.exists('样品登记单.xlsx'):
    online_reg_file = openpyxl.load_workbook('样品登记单.xlsx')
    online_reg_info_sheet = online_reg_file['分析基本信息']
    online_sample_info_sheet = online_reg_file['样本信息']
    online_group_info_sheet = online_reg_file['样本分组信息']
    if os.path.exists('项目登记单.xlsx'):
        # 加载项目登记文件
        project_reg_file = openpyxl.load_workbook('项目登记单.xlsx')
        project_reg_sheet = project_reg_file['项目登记单']
        project_sample_info_sheet = project_reg_file['样本信息']
        project_group_info_sheet = project_reg_file['分组信息']
        project_compare_info_sheet = project_reg_file['比较组信息']

        # 从线上登记表中获取值
        project_no_value = get_value_from_sheet(online_reg_info_sheet, '项目编号')
        customer_name_value = get_value_from_sheet(online_reg_info_sheet, '客户名称')
        contact_person_value = get_value_from_sheet(online_reg_info_sheet, '联系人')
        sample_species_value = get_value_from_sheet(online_reg_info_sheet, '样本物种')
        sample_type_value = get_value_from_sheet(online_reg_info_sheet, '样本类型')

        # 在项目登记表中更新值
        value_mapping = {
            '项目编号': project_no_value,
            '客户名称': customer_name_value,
            '联系人名称': contact_person_value,
            '物种': sample_species_value,
            '样本类型': sample_type_value,
            '执行编码': execution_code
        }

        for row in range(1, project_reg_sheet.max_row + 1):
            cell_value = project_reg_sheet.cell(row=row, column=1).value
            if cell_value in value_mapping:
                project_reg_sheet.cell(row=row, column=2).value = value_mapping[cell_value]

        # 在项目登记表中填充分组信息
        remove_info(project_group_info_sheet)
        copy_info(online_group_info_sheet, project_group_info_sheet)
        remove_info(project_sample_info_sheet)
        copy_info(online_group_info_sheet, project_sample_info_sheet)

        # 在分组信息和比较组信息表中填充其他列

        raw_folder_path = 'raw'
        # 检查 raw 文件夹是否存在
        if os.path.exists(raw_folder_path) and os.path.isdir(raw_folder_path):
            # 判断“neg”文件夹是否存在
            if os.path.exists(raw_folder_path + '/neg') and os.path.isdir(raw_folder_path + '/neg'):
                # 判断“pos”文件夹是否存在
                if os.path.exists(raw_folder_path + '/pos') and os.path.isdir(raw_folder_path + '/pos'):
                    mode_value = 'both'
                else:
                    mode_value = 'neg'
            # 判断“pos”文件夹是否存在
            elif os.path.exists(raw_folder_path + '/pos') and os.path.isdir(raw_folder_path + '/pos'):
                mode_value = 'pos'
            else:
                mode_value = ''  # 如果都不存在，则模式为空
        else:
            mode_value = ''  # 如果raw文件夹不存在，则模式为空

        # 将判断得到的模式值填入“项目登记单.xlsx”表格中的“分组信息”工作表的“模式”列
        for row in range(2, project_group_info_sheet.max_row + 1):
            project_group_info_sheet.cell(row=row, column=3).value = 'ALL'
            project_group_info_sheet.cell(row=row, column=4).value = mode_value

        # “项目登记单.xlsx”表格中的"比较组信息“工作表信息填写
        if '差异比较信息' in online_reg_file.sheetnames:
            online_diff_info_sheet = online_reg_file['差异比较信息']
            # 获取在线差异比较表中“case”和“control”列的索引
            online_headers = [cell.value for cell in online_diff_info_sheet[1]]
            case_column_index = online_headers.index('case')
            control_column_index = online_headers.index('control')
            for row_index in range(2, online_diff_info_sheet.max_row + 1):
                case_value = online_diff_info_sheet.cell(row=row_index, column=case_column_index + 1).value
                control_value = online_diff_info_sheet.cell(row=row_index, column=control_column_index + 1).value
                project_compare_info_sheet.cell(row=row_index, column=1).value = case_value
                project_compare_info_sheet.cell(row=row_index, column=2).value = control_value
                project_compare_info_sheet.cell(row=row_index, column=3).value = case_value+"/"+control_value
                project_compare_info_sheet.cell(row=row_index, column=4).value = mode_value
                for online_group_row in online_group_info_sheet.iter_rows(min_row=2, values_only=True):
                    group_name = online_group_row[0]
                    sample_names = online_group_row[1]
                    if case_value in group_name:
                        if ',' in sample_names:
                            project_compare_info_sheet.cell(row=row_index, column=5).value = 'both'
                        else:
                            project_compare_info_sheet.cell(row=row_index, column=5).value = 'all'
                        break
        else:
            print("不存在比较组信息")

        # 保存修改后的项目登记文件
        project_reg_file.save('项目登记单.xlsx')
    else:
        print("当前目录不存在项目登记单")
else:
    print("当前目录不存在样品登记单")
